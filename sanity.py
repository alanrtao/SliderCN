import sys
from os import path, listdir
import csv

# Configurations

font_paths = ('fusion-pixel-12px-proportional-latin.ttf', 'fusion-pixel-12px-proportional-zh_hans.ttf')
err_log = './errors.log'

parent = path.realpath(path.dirname(__file__))
header_row_count = 4 # format comments, config name, config val, column names
ban_characters = '。：；，'

locale_name = '简体中文'

if len(sys.argv) < 2:
    print('Usage:')
    print(f'  - `python sanity.py split` splits `./{locale_name}.xlsx` into separate csv under ./{locale_name}/<sheet>.csv, which will override existing .csv value')
    print('  - `python sanity.py combine does the opposite, which will override existing .xlsx value.`')
    print('  - `python sanity.py scan` checks the CSVs for errors, generating an error log as well as a usage log')
    sys.exit(1)

mode = sys.argv[1]

csv_paths = {f[:-4]: path.join(parent, f) for f in listdir(parent) if f.endswith('.csv')}

#########
# Split #
#########
if mode == 'split':
    from openpyxl import load_workbook
    from io import StringIO

    wb = load_workbook(path.realpath(f'./{locale_name}.xlsx'), read_only=True, rich_text=False)

    for s_name in wb.sheetnames:
        print(s_name)
        s = wb[s_name]
        out = StringIO(newline='')
        w = csv.writer(out)

        for row in s.rows:
            row_out = []
            for col in row:
                if col.value is None:
                    row_out.append('')
                else:
                    row_out.append(str(col.value))
            w.writerow(row_out)

        with open(path.realpath(f'./{s_name}.csv'), mode='w') as f:
            f.write(out.getvalue())
        
        

###########
# Combine #
###########
elif mode == 'combine':
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    w = Workbook(write_only=True)

    for i, (csv_name, csv_path) in enumerate(sorted(csv_paths.items(), key=lambda x: x[0])):
        print(i, csv_name, csv_path)
        ws: Worksheet = w.create_sheet(csv_name, index=i)
        with open(csv_path) as f:
            r = csv.reader(f)
            for row in r:
                ws.append(row)

    w.save(f'{locale_name}.xlsx')

########
# SCAN #
########
elif mode == 'scan':
    from textwrap import indent
    from fontTools.ttLib import TTFont
    from dataclasses import dataclass
    import re
        
    @dataclass
    class Record:
        path: str
        orig: str
        translation: str
        metadata: str

    @dataclass
    class FontDescriptor:
        path: str
        usage: set[str]

    fonts: dict[TTFont, FontDescriptor] = { TTFont(p): FontDescriptor(p, set()) for p in font_paths }

    def char_in_font(c: str, f: TTFont) -> bool:
        cmaps = f['cmap'].tables
        for cm in cmaps:
            # if cm.isUnicode():
            if ord(c) in cm.cmap:
                fonts[f].usage.add(c)
                return True
        return False

    def char_not_in_fonts(c: str) -> bool:
        for f in fonts.keys():
            if char_in_font(c, f): return False
        return 
    
    def get_tags(s: str) -> list[str]:
        tags = re.findall(r'<.*?>', s)
        return sorted(tags)

    with open(err_log, mode='w') as log_file:
        def printdup(*args, **kwargs) -> None:
            print(*args, **kwargs)
            print(*args, **kwargs, file=log_file)

        def log_error(msg: str, lineno: int, rec: Record):
            printdup(msg)
            printdup(lineno)
            printdup('Original:')
            printdup(indent(rec.orig, '  '))
            printdup('Translation:')
            printdup(indent(rec.translation, '  '))

        cmaps = []

        for csv_path in sorted(csv_paths.values()):
            with open(csv_path) as csv_file:
                printdup(f'=====================')
                printdup(f'BEGIN {csv_path}')
                printdup(f'=====================')

                r = csv.reader(csv_file)

                for i in range(header_row_count):
                    next(r)
                
                for i, row in enumerate(r):
                    rec = Record(*(row[:4]))

                    def log_curr_error(msg: str): 
                        log_error(msg, i, rec)

                    if rec.translation != '' and rec.translation.lstrip() == '':
                        log_curr_error('Nonempty whitespace line.')
                        continue

                    if rec.translation == '' and rec.orig != '':
                        log_curr_error('Missing translation (translation column empty).')
                        continue

                    if re.search(f'[a-zA-Z]', rec.orig) is not None and rec.orig == rec.translation:
                        log_curr_error('Missing translation (translation column autofilled).')
                    
                    for bc in ban_characters:
                        if bc in rec.translation:
                            log_curr_error(f'Character explicitly forbidden: `{bc}`')
                    
                    for c in rec.translation:
                        # skip whitespaces
                        if c.lstrip() == '':
                            continue
                        if char_not_in_fonts(c):
                            log_curr_error(f'Character not supported by font `{c}`')

                    tags = get_tags(rec.orig)
                    tags_ = get_tags(rec.translation)
                    if tags != tags_:
                        log_curr_error(f'Difference between tags in orig and translation\nOriginal tags: {tags}\nTranslation tags: {tags_}')


                
                printdup(f'=====================')
                printdup(f'END {csv_path}')
                printdup(f'=====================')
            
    for font, desc in fonts.items():
        with open(f'{desc.path}.log', mode='w') as w:
            for u in sorted(desc.usage):
                print(u, end='', file=w)

else:
    raise ValueError(f'{mode} does not fall under `split`, `fall`, `scan`')

